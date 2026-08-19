"""Format loaders: one loader per file type, each producing the SAME
`RecoveredDocument` so the DOM builder stays format-agnostic.

Extensibility rule (SYN4): adding DICOM/CAD/etc. means adding a loader here
that returns RecoveredDocument — nothing downstream changes.

Design notes:
  * read the file ONCE (single extraction pass).
  * OCR (scanned + images) is a loader backend, invoked only for image/scanned.
  * tables and images are first-class parts; never flattened to prose.
"""
from __future__ import annotations

import csv as _csv
import hashlib
import html.parser
import io
import json
from xml.etree import ElementTree as ET

from ..config import ParserConfig
from ..mime import MIME as _MIME
from ..parts import (
    RecoveredBlock,
    RecoveredDocument,
    RecoveredImage,
    RecoveredTable,
)


class UnsupportedFormat(Exception):
    """Raised when a detected type has no loader (surfaced as unsupported)."""


class _TextExtractor(html.parser.HTMLParser):
    """Minimal block-aware HTML -> [RecoveredBlock], without any deps."""

    _BLOCK = {"p", "h1", "h2", "h3", "h4", "h5", "h6", "li", "pre", "div", "td", "th"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self._cur = {"tag": None, "text": []}
        self.blocks = []

    def handle_starttag(self, tag, attrs):
        if tag in self._BLOCK:
            self.flush()
            self._cur["tag"] = tag
        if tag == "br":
            self._cur["text"].append("\n")
        if tag == "a":
            self.link_depth = getattr(self, "link_depth", 0) + 1
            self.linktext = []

    def handle_endtag(self, tag):
        if tag in self._BLOCK:
            self.flush()
        if tag == "a":
            self.link_depth = getattr(self, "link_depth", 0) - 1

    def handle_data(self, data):
        if self.link_depth > 0:
            self.linktext.append(data)
        self._cur["text"].append(data)

    def flush(self):
        text = "".join(self._cur["text"]).strip()
        if text:
            tag = self._cur["tag"] or "p"
            kind = "heading" if tag[0] == "h" else ("list_item" if tag == "li" else "paragraph")
            self.blocks.append((kind, text))
        self._cur = {"tag": None, "text": []}


class Loaders:
    def __init__(self, config: ParserConfig):
        self.config = config

    def load(self, detected, data: bytes, *, route: str | None = None) -> RecoveredDocument:
        slug = detected.slug
        # ADR-007 + ADR-011 (route-aware dispatch). When a `route` is supplied
        # (from the Extractor after detection) we honor it; when it's None we
        # fall back to the legacy behaviour decided purely by
        # `config.layout_backend` ("docling"/"native" manual overrides, which
        # are preserved). In "auto", a non-PDF gets route=None here and stays on
        # its existing native loader path (Gap A: images keep native OCR).
        if route is None:
            route = self.config.layout_backend

        # docling tier: PDFs (and bare images under a manual "docling" override)
        if route == "docling" and slug in ("pdf", "png", "jpg", "gif", "tiff"):
            from . import docling_loader

            if docling_loader.engine_available():
                rec = docling_loader.parse(data, f"doc.{slug}", self.config.docling_models_dir)
                if rec is not None:
                    rec.detected_type = slug
                    rec.mime = detected.mime
                    rec.declared_extension = detected.declared_extension
                    rec.probe = detected.probe
                    return rec

        # enrichment tier (auto-routed PDFs): native extraction + OCR post-pass
        if route == "enrichment" and slug == "pdf":
            rec = self._pdf(data, detected)
            from . import enrichment

            rec = enrichment.enrich_scanned_pages(rec, self.config, data=data)
            rec.reading_order_authoritative = False
            return rec

        # native per-format paths unchanged (covers route in {"auto","native",
        #  None} for PDFs and every non-PDF format).
        if slug in ("plaintext", "txt"):
            return self._text(data, detected)
        if slug in ("png", "jpg", "gif", "tiff"):
            return self._image(data, detected)
        if slug == "pdf":
            return self._pdf(data, detected)
        if slug == "docx":
            return self._docx(data, detected)
        if slug == "xlsx":
            return self._xlsx(data, detected)
        if slug == "csv":
            return self._delimited(data, detected, ",")
        if slug == "tsv":
            return self._delimited(data, detected, "\t")
        if slug == "json":
            return self._json(data, detected)
        if slug == "xml":
            return self._xml(data, detected)
        if slug == "html":
            return self._html(data, detected)
        if slug in ("markdown", "md"):
            return self._markdown(data, detected)
        raise UnsupportedFormat(slug)

    # --- metadata shim -----------------------------------------------------
    def _base(self, detected, RecoveredDocument):
        return RecoveredDocument(
            detected_type=detected.slug,
            mime=detected.mime,
            declared_extension=detected.declared_extension,
            probe=detected.probe,
        )

    # --- PDF ---------------------------------------------------------------
    def _pdf(self, data, detected):
        import fitz  # PyMuPDF

        try:
            doc = fitz.open(stream=data, filetype="pdf")
        except Exception as e:
            raise UnsupportedFormat(f"pdf:{e}")
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = doc.page_count

        from ._pdfmeta import fitz_metadata

        for key, val in fitz_metadata(doc).items():
            setattr(rec, key, val)

        # Delegate to the single-source-of-truth per-page native extractor
        # (ADR-013 T5). The whole doc is already open from bytes here; each page
        # is folded into one RecoveredDocument — no duplicated extraction logic.
        from ..engines.native_pdf import _native_page_from_doc

        seq = 0
        # F1: document-wide median body size for heading classification, matching
        # the legacy behaviour (a block is a heading if its font size exceeds the
        # whole-document median * threshold). Computed once, then shared.
        _sizes: list[float] = []
        for pno in range(doc.page_count):
            for blk in doc[pno].get_text("dict").get("blocks", []):
                if blk.get("type") != 0:
                    continue
                for line in blk.get("lines", []):
                    for span in line.get("spans", []):
                        if span.get("text", "").strip():
                            _sizes.append(float(span.get("size", 0.0)))
        _body_med = sorted(_sizes)[len(_sizes) // 2] if _sizes else 12.0
        for pno in range(doc.page_count):
            rec.page_sizes[pno] = (doc[pno].rect.width, doc[pno].rect.height)
            pr = _native_page_from_doc(doc[pno], pno, self.config, body_med=_body_med)
            for b in pr.blocks:
                b.seq = seq
                seq += 1
                rec.blocks.append(b)
            rec.tables.extend(pr.tables)
            rec.images.extend(pr.images)
        return rec

    # --- CSV / TSV -----------------------------------------------------------
    def _delimited(self, data, detected, delimiter):
        rec = self._base(detected, RecoveredDocument)
        text = data.decode("utf-8-sig", errors="replace")
        reader = list(_csv.reader(io.StringIO(text), delimiter=delimiter))
        rows = [r for r in reader if any(c.strip() for c in r)]
        if not rows:
            return rec
        rec.page_count = 1
        header = [c.strip() for c in rows[0]]
        rec.tables.append(
            RecoveredTable(page=0, header=header, rows=[[c.strip() for c in r] for r in rows[1:]], source="native")
        )
        return rec

    # --- simple text ---------------------------------------------------------
    def _text(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        text = data.decode("utf-8-sig", errors="replace")
        rec.page_count = 1
        for i, line in enumerate(text.split("\n")):
            if line.strip():
                rec.blocks.append(RecoveredBlock(page=0, seq=i, text=line.strip(), kind="paragraph", source="text"))
        return rec

    def _json(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = 1
        try:
            obj = json.loads(data.decode("utf-8-sig", errors="ignore"))
        except Exception:
            obj = None
        rec.blocks.append(
            RecoveredBlock(page=0, seq=0, text=json.dumps(obj, ensure_ascii=False) if obj is not None else "", kind="code", source="json")
        )
        return rec

    def _xml(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = 1
        try:
            root = ET.fromstring(data)
        except Exception:
            root = None
        text = ET.tostring(root, encoding="unicode") if root is not None else data.decode("utf-8", errors="ignore")
        rec.blocks.append(RecoveredBlock(page=0, seq=0, text=text, kind="code", source="xml"))
        return rec

    def _html(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = 1
        p = _TextExtractor()
        try:
            p.feed(data.decode("utf-8-sig", errors="replace"))
        except Exception:
            pass
        p.close()
        for i, (kind, text) in enumerate(p.blocks):
            rec.blocks.append(RecoveredBlock(page=0, seq=i, kind=kind, text=text, source="markup"))
        return rec

    def _markdown(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = 1
        lines = data.decode("utf-8-sig", errors="replace").splitlines()
        para = []
        i = 0
        for raw in lines:
            line = raw.rstrip()
            if not line.strip():
                if para:
                    rec.blocks.append(RecoveredBlock(page=0, seq=len(rec.blocks), text=" ".join(para).strip(), source="markdown"))
                    para = []
                continue
            if line.lstrip().startswith(("#", "```", "-", "*")):
                if para:
                    rec.blocks.append(RecoveredBlock(page=0, seq=len(rec.blocks), text=" ".join(para).strip(), source="markdown"))
                    para = []
                if line.lstrip().startswith("#"):
                    kind = "heading"
                elif line.lstrip().startswith("```"):
                    kind = "code"
                    para.append(line)
                    continue
                else:
                    kind = "list_item"
                rec.blocks.append(RecoveredBlock(page=0, seq=len(rec.blocks), text=line.strip(), kind=kind, source="markdown"))
            else:
                para.append(line)
        if para:
            rec.blocks.append(RecoveredBlock(page=0, seq=len(rec.blocks), text=" ".join(para).strip(), source="markdown"))
        return rec

    # --- DOCX ---------------------------------------------------------------
    def _docx(self, data, detected):
        import zipfile
        rec = self._base(detected, RecoveredDocument)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
            root = ET.fromstring(zf.read("word/document.xml"))
            core = zf.read("word/core.xml") if "word/core.xml" in zf.namelist() else b""
        except Exception as e:
            raise UnsupportedFormat(f"docx:{e}")
        body = root.find(".//w:body", ns)
        rec.page_count = 1
        seq = 0
        if body is not None:
            for child in body:
                if child.tag == "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p":
                    text = "".join(t.text or "" for t in child.iter(_NSW("t")))
                    if text.strip():
                        rec.blocks.append(RecoveredBlock(page=0, seq=seq, text=text.strip(), source="markup"))
                        seq += 1
                elif child.tag == _NSW("tbl"):
                    rows = []
                    for tr in child.findall(_NSW("tr")):
                        row = []
                        for tc in tr.findall(_NSW("tc")):
                            row.append("".join(t.text or "" for t in tc.iter(_NSW("t"))).strip())
                        if any(r for r in row):
                            rows.append(row)
                    if rows:
                        rec.tables.append(RecoveredTable(page=0, header=rows[0], rows=rows[1:], source="native"))
        # core props
        if core:
            cns = {"cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"}
            try:
                croot = ET.fromstring(core)
                rec.title = (croot.findtext(".//cp:title", None, cns) or "").strip()
                rec.creator = (croot.findtext(".//cp:creator", None, cns) or "").strip()
            except Exception:
                pass
        return rec

    # --- XLSX ---------------------------------------------------------------
    def _xlsx(self, data, detected):
        from openpyxl import load_workbook
        rec = self._base(detected, RecoveredDocument)
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        rec.page_count = len(wb.sheetnames)
        for sid, ws in enumerate(wb.worksheets):
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([("" if v is None else str(v)) for v in row])
            rows = [r for r in rows if any(c.strip() for c in r)]
            if not rows:
                continue
            rec.tables.append(
                RecoveredTable(page=sid, header=[c.strip() for c in rows[0]],
                               rows=[[c.strip() for c in r] for r in rows[1:]], source="native")
            )
            rec.page_sizes[sid] = (ws.max_column or 0, ws.max_row or 0)
        return rec

    # --- image / scanned -----------------------------------------------------
    def _image(self, data, detected):
        rec = self._base(detected, RecoveredDocument)
        rec.page_count = 1
        _image_bytes(rec, data, self.config)
        return rec


def _NSW(tag):
    return "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}" + tag


def _image_bytes(rec: RecoveredDocument, data: bytes, config: ParserConfig) -> None:
    """Shared image→OCR helper used by both `Loaders._image` and the standalone
    `ImageEngine`. Appends OCR `RecoveredBlock`s (page 0) to `rec` in place.

    OCR is gated by `config.ocr_enabled`; with it disabled the record stays
    empty but OK (no crash). Reuses the on-prem `ocr.ocr_bytes` wrapper.
    """
    if not config.ocr_enabled:
        return
    import time as _time
    from .. import ocr

    t_ocr = _time.time()
    lines = ocr.ocr_bytes(data)
    rec.timings["ocr_ms"] = round((_time.time() - t_ocr) * 1000, 1)
    rec.timings["ocr_pages"] = 1
    for i, (text, bbox, conf) in enumerate(lines):
        rec.blocks.append(
            RecoveredBlock(page=0, seq=i, text=text, bbox=bbox,
                           confidence=conf if conf <= 1.0 else conf / 100.0,
                           source="ocr", ocr_engine=ocr.engine_name())
        )